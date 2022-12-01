import datetime
import math
from pathlib import Path

import numpy
import os
import openpyxl
import xlsxwriter
from random import *
from openpyxl.utils import get_column_letter
from deftech import tech, grouptechs, learnTech


def ResetGame():
    NewGamePlayer('crono')
    NewGamePlayer('lucca')
    NewGamePlayer('marle')
    NewGamePlayer('frog')
    NewGamePlayer('robo')
    NewGamePlayer('ayla')
    NewGamePlayer('magus')


def colNameToNum(name):
    pow1 = 1
    colNum1 = 0
    for letter in name[::-1]:
        colNum1 += (int(letter, 36) - 9) * pow1
        pow1 *= 26
    return colNum1


def NewGamePlayer(Player):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1

    p = Player.lower()
    print('[Process:', localvar, ']', 'NewGamePlayer', p)

    if p == "crono":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Wooden Sword', 'crono')
        EquipItem('Helm', 'Hide Cap', 'crono')
        EquipItem('Armor', 'Hide Tunic', 'crono')
        EquipItem('Accessory', 'Headband', 'crono')
    elif p == "marle":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Bronze Bowgun', 'marle')
        EquipItem('Helm', 'Hide Cap', 'marle')
        EquipItem('Armor', 'Hide Tunic', 'marle')
        EquipItem('Accessory', 'Ribbon', 'marle')
    elif p == "lucca":
        LevelUpPlayer(p, 2)
        EquipItem('Weapon', 'Airgun', 'lucca')
        EquipItem('Helm', 'Hide Cap', 'lucca')
        EquipItem('Armor', 'Hide Tunic', 'lucca')
        EquipItem('Accessory', 'Sight Scope', 'lucca')
    elif p == "frog":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Bronze Sword', 'frog')
        EquipItem('Helm', 'Bronze Helm', 'frog')
        EquipItem('Armor', 'Bronze Armor', 'frog')
        EquipItem('Accessory', 'PowerGlove', 'frog')
    elif p == "robo":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Tin Arm', 'robo')
        EquipItem('Helm', 'Iron Helm', 'robo')
        EquipItem('Armor', 'Titanium Vest', 'robo')
        EquipItem('Accessory', 'Guardian Bangle', 'robo')
    elif p == "ayla":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Fist I', 'ayla')
        EquipItem('Helm', 'Stone Helm', 'ayla')
        EquipItem('Armor', 'Ruby Vest', 'ayla')
        EquipItem('Accessory', 'Power Scarf', 'ayla')
    elif p == "magus":
        LevelUpPlayer(p, 1)
        EquipItem('Weapon', 'Moonfall Scythe', 'magus')
        EquipItem('Helm', 'Doom Helm', 'magus')
        EquipItem('Armor', 'Raven Armor', 'magus')
        EquipItem('Accessory', "Schala's Amulet", 'magus')

    print('[Process:', localvar, ']', 'Success')
    return 0


def ReadStatus(Player):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'a')
    print('[Process:', localvar, ']', 'Success')


def Inventory(action, name=None, nametype='item', mod=None):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    # print('[Process:', localvar, ']', 'Inventory', action, name, nametype, mod)
    # Action: remove (-1 of given item), add (+1 of given item), check (verify if quantity >= 1, list (list all items)
    # Return:      1 if success              1 if success,           1 if success,           list

    Inv = Path("Inventory.xlsx")
    Invobj = openpyxl.load_workbook(Inv, data_only=True)
    LotofThings = Invobj.active
    maxrow = LotofThings.max_row + 1

    tempCDE = ['C', 'D', 'E']

    if action.lower() == 'list':
        LotofItems = [['' for x in range(3)] for y in range(maxrow)]
        for i in range(3, maxrow):
            if LotofThings['C' + str(i)].value is None:
                break
            for j in range(3):
                # print(i, j, LotofThings['C' + str(i)].value)
                LotofItems[i - 3][j] = LotofThings[tempCDE[j] + str(i)].value
        # print(LotofItems)

        # print('[Process:', localvar, ']', 'Success')
        return LotofItems
    elif (action.lower() == 'add' or action.lower() == 'remove') and name is not None:
        # print('plus')

        if nametype.lower() == 'gold':
            LotofThings['H2'] = int(LotofThings['H2'].value) + int(name)
            Invobj.save(Inv)
            return 0
        invlist = Inventory('list')
        continuei = False
        iii = 0
        for i in range(maxrow):
            if name.lower() == str(invlist[i][2]).lower():
                # print('plus2')
                # print(tempCDE[1], i, LotofThings[tempCDE[1] + str(i+3)].value)
                if action.lower() == 'add':
                    addtofile = 1
                else:
                    addtofile = -1
                LotofThings[tempCDE[1] + str(i + 3)] = int(LotofThings[tempCDE[1] + str(i + 3)].value) + addtofile
                if LotofThings[tempCDE[1] + str(i + 3)].value <= 0:
                    LotofThings[tempCDE[1] + str(i + 3)] = 0
                continuei = True
                break
        if continuei is False:
            # print(maxrow, i, name)
            if action.lower() == 'add':
                addtofile = 1
            else:
                # print('[Process:', localvar, ']', 'Success')
                return 0
            LotofThings['C' + str(maxrow)] = nametype
            LotofThings['D' + str(maxrow)] = addtofile
            LotofThings['E' + str(maxrow)] = name
    elif action.lower() == 'check' and name is not None:
        invlist = Inventory('list')
        for i in range(len(invlist)):
            if invlist[i][2].lower() == name.lower():
                if int(invlist[i][1]) > 0:
                    # print('[Process:', localvar, ']', 'Success')
                    return True
        # print('[Process:', localvar, ']', 'Success')
        return False
    else:
        # print('[Process:', localvar, ']', 'Success')
        return 0
    Invobj.save(Inv)

    # print('[Process:', localvar, ']', 'Success')


def ReadBossData(name):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'ReadMonsterData', name)

    Bosses = Path("Bosses.xlsx")
    Bossesobj = openpyxl.load_workbook(Bosses, data_only=True)
    Bossesxlsx = Bossesobj.active

    EnemyStatus = [['' for x in range(1)] for y in range(19)]
    name = name.lower()
    test = False
    # Return the vector EnemyStatus[19] := HP, XP, TP, GOLD, Charmables[255], Drops[255], Location[255], ATK, DEF,
    # MAG, MDEF, SPD, EVA, HIT, Weakness[255], Resis[255], Absorbs[255], Techs[255], Counter[255].
    for i in range(2048):
        temprp = colNameToNum('C')
        i = 8 * i
        temp = Bossesxlsx[str(get_column_letter(temprp + i)) + '4'].value
        # print(temp, name)
        if temp is None:
            break
        if temp.lower() == name:
            test = True
            # print('debug', name, get_column_letter(temprp+i)+'4')
            line1 = 0
            line2 = 0
            line3 = 0
            # print(EnemyStatus)
            for j in range(255):
                if Bossesxlsx[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'HP':
                    line1 = str(5 + j + 1)
                elif Bossesxlsx[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'Attack':
                    line2 = str(5 + j + 1)
                elif Bossesxlsx[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'Weaknesses':
                    line3 = str(5 + j + 1)
            EnemyStatus[0][0] = Bossesxlsx[str(get_column_letter(temprp + i + 1)) + line1].value  # HP
            EnemyStatus[1][0] = Bossesxlsx[str(get_column_letter(temprp + i + 2)) + line1].value  # XP
            EnemyStatus[2][0] = Bossesxlsx[str(get_column_letter(temprp + i + 3)) + line1].value  # TP
            EnemyStatus[3][0] = Bossesxlsx[str(get_column_letter(temprp + i + 4)) + line1].value  # GOLD

            EnemyStatus[4][0] = Bossesxlsx[str(get_column_letter(temprp + i + 5)) + line1].value  # CHARMABLES
            if EnemyStatus[4][0] != "None":
                # print('test', EnemyStatus[4][0])
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 5)) + str(int(line1) + j + 1)].value
                    if tempes != 'Speed':
                        EnemyStatus[4].append(tempes)
                    else:
                        break

            EnemyStatus[5][0] = Bossesxlsx[str(get_column_letter(temprp + i + 6)) + line1].value  # DROPS
            if EnemyStatus[5][0] != "None":
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 6)) + str(int(line1) + j + 1)].value
                    if tempes != 'Evade':
                        EnemyStatus[5].append(tempes)
                    else:
                        break

            EnemyStatus[6][0] = Bossesxlsx[str(get_column_letter(temprp + i + 7)) + line1].value  # LOCATION
            if EnemyStatus[6][0] != "None":
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 7)) + str(int(line1) + j + 1)].value
                    if tempes != 'Hit':
                        EnemyStatus[6].append(tempes)
                    else:
                        break

            EnemyStatus[7][0] = Bossesxlsx[str(get_column_letter(temprp + i + 1)) + line2].value  # ATK
            EnemyStatus[8][0] = Bossesxlsx[str(get_column_letter(temprp + i + 2)) + line2].value  # DEF
            EnemyStatus[9][0] = Bossesxlsx[str(get_column_letter(temprp + i + 3)) + line2].value  # MAG
            EnemyStatus[10][0] = Bossesxlsx[str(get_column_letter(temprp + i + 4)) + line2].value  # MDEF
            EnemyStatus[11][0] = Bossesxlsx[str(get_column_letter(temprp + i + 5)) + line2].value  # SPD
            EnemyStatus[12][0] = Bossesxlsx[str(get_column_letter(temprp + i + 6)) + line2].value  # EVA
            EnemyStatus[13][0] = Bossesxlsx[str(get_column_letter(temprp + i + 7)) + line2].value  # HIT

            EnemyStatus[14][0] = Bossesxlsx[str(get_column_letter(temprp + i + 1)) + line3].value  # WEAKNESS
            if EnemyStatus[14][0] != "None":
                # print(EnemyStatus[14][0])
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 1)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[14].append(tempes)
                    else:
                        break

            EnemyStatus[15][0] = Bossesxlsx[str(get_column_letter(temprp + i + 2)) + line3].value  # RESIS
            if EnemyStatus[15][0] != "None":
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 2)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[15].append(tempes)
                    else:
                        break

            EnemyStatus[16][0] = Bossesxlsx[str(get_column_letter(temprp + i + 3)) + line3].value  # ABSORBS
            if EnemyStatus[16][0] != "None":
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 3)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[16].append(tempes)
                    else:
                        break

            EnemyStatus[17][0] = Bossesxlsx[str(get_column_letter(temprp + i + 4)) + line3].value  # TECHS
            if EnemyStatus[17][0] != "None":
                for j in range(255):
                    # print(get_column_letter((temprp+i+4)), int(line3)+j+1)
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 4)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[17].append(tempes)
                    else:
                        break

            EnemyStatus[18][0] = Bossesxlsx[str(get_column_letter(temprp + i + 6)) + line3].value  # COUNTER
            if EnemyStatus[18][0] != "None":
                for j in range(255):
                    tempes = Bossesxlsx[str(get_column_letter(temprp + i + 6)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[18].append(tempes)
                    else:
                        break

            # print('debug', line1, line2, line3)
            # print('debug', name, EnemyStatus)
            wait = input()
            break
        i = i / 8

    # Enemiesobj.save(Enemies)
    # print('[Process:', localvar, ']', 'Success')
    if test:
        return EnemyStatus
    else:
        return None


def ReadMonsterData(name):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'ReadMonsterData', name)

    Enemies = Path("Enemies.xlsx")
    Enemiesobj = openpyxl.load_workbook(Enemies, data_only=True)
    Ene = Enemiesobj.active

    EnemyStatus = [['' for x in range(1)] for y in range(19)]
    name = name.lower()
    test = False
    # Return the vector EnemyStatus[19] := HP, XP, TP, GOLD, Charmables[255], Drops[255], Location[255], ATK, DEF,
    # MAG, MDEF, SPD, EVA, HIT, Weakness[255], Resis[255], Absorbs[255], Techs[255], Counter[255].
    for i in range(2048):
        temprp = colNameToNum('C')
        i = 8 * i
        temp = Ene[str(get_column_letter(temprp + i)) + '4'].value
        # print(temp, name)
        if temp is None:
            break
        if temp.lower() == name:
            test = True
            # print('debug', name, get_column_letter(temprp+i)+'4')
            line1 = 0
            line2 = 0
            line3 = 0
            # print(EnemyStatus)
            for j in range(255):
                if Ene[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'HP':
                    line1 = str(5 + j + 1)
                elif Ene[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'Attack':
                    line2 = str(5 + j + 1)
                elif Ene[str(get_column_letter(temprp + i + 1)) + str(5 + j)].value == 'Weaknesses':
                    line3 = str(5 + j + 1)
            EnemyStatus[0][0] = Ene[str(get_column_letter(temprp + i + 1)) + line1].value  # HP
            EnemyStatus[1][0] = Ene[str(get_column_letter(temprp + i + 2)) + line1].value  # XP
            EnemyStatus[2][0] = Ene[str(get_column_letter(temprp + i + 3)) + line1].value  # TP
            EnemyStatus[3][0] = Ene[str(get_column_letter(temprp + i + 4)) + line1].value  # GOLD

            EnemyStatus[4][0] = Ene[str(get_column_letter(temprp + i + 5)) + line1].value  # CHARMABLES
            if EnemyStatus[4][0] != "None":
                # print('test', EnemyStatus[4][0])
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 5)) + str(int(line1) + j + 1)].value
                    if tempes != 'Speed':
                        EnemyStatus[4].append(tempes)
                    else:
                        break

            EnemyStatus[5][0] = Ene[str(get_column_letter(temprp + i + 6)) + line1].value  # DROPS
            if EnemyStatus[5][0] != "None":
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 6)) + str(int(line1) + j + 1)].value
                    if tempes != 'Evade':
                        EnemyStatus[5].append(tempes)
                    else:
                        break

            EnemyStatus[6][0] = Ene[str(get_column_letter(temprp + i + 7)) + line1].value  # LOCATION
            if EnemyStatus[6][0] != "None":
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 7)) + str(int(line1) + j + 1)].value
                    if tempes != 'Hit':
                        EnemyStatus[6].append(tempes)
                    else:
                        break

            EnemyStatus[7][0] = Ene[str(get_column_letter(temprp + i + 1)) + line2].value  # ATK
            EnemyStatus[8][0] = Ene[str(get_column_letter(temprp + i + 2)) + line2].value  # DEF
            EnemyStatus[9][0] = Ene[str(get_column_letter(temprp + i + 3)) + line2].value  # MAG
            EnemyStatus[10][0] = Ene[str(get_column_letter(temprp + i + 4)) + line2].value  # MDEF
            EnemyStatus[11][0] = Ene[str(get_column_letter(temprp + i + 5)) + line2].value  # SPD
            EnemyStatus[12][0] = Ene[str(get_column_letter(temprp + i + 6)) + line2].value  # EVA
            EnemyStatus[13][0] = Ene[str(get_column_letter(temprp + i + 7)) + line2].value  # HIT

            EnemyStatus[14][0] = Ene[str(get_column_letter(temprp + i + 1)) + line3].value  # WEAKNESS
            if EnemyStatus[14][0] != "None":
                # print(EnemyStatus[14][0])
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 1)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[14].append(tempes)
                    else:
                        break

            EnemyStatus[15][0] = Ene[str(get_column_letter(temprp + i + 2)) + line3].value  # RESIS
            if EnemyStatus[15][0] != "None":
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 2)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[15].append(tempes)
                    else:
                        break

            EnemyStatus[16][0] = Ene[str(get_column_letter(temprp + i + 3)) + line3].value  # ABSORBS
            if EnemyStatus[16][0] != "None":
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 3)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[16].append(tempes)
                    else:
                        break

            EnemyStatus[17][0] = Ene[str(get_column_letter(temprp + i + 4)) + line3].value  # TECHS
            if EnemyStatus[17][0] != "None":
                for j in range(255):
                    # print(get_column_letter((temprp+i+4)), int(line3)+j+1)
                    tempes = Ene[str(get_column_letter(temprp + i + 4)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[17].append(tempes)
                    else:
                        break

            EnemyStatus[18][0] = Ene[str(get_column_letter(temprp + i + 6)) + line3].value  # COUNTER
            if EnemyStatus[18][0] != "None":
                for j in range(255):
                    tempes = Ene[str(get_column_letter(temprp + i + 6)) + str(int(line3) + j + 1)].value
                    if tempes is not None:
                        EnemyStatus[18].append(tempes)
                    else:
                        break

            # print('debug', line1, line2, line3)
            # print('debug', name, EnemyStatus)
            break
        i = i / 8

    # Enemiesobj.save(Enemies)
    # print('[Process:', localvar, ']', 'Success')
    if test:
        return EnemyStatus
    else:
        return None


def ReadPlayerData(name):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    # print('[Process:', localvar, ']', 'ReadPartyData', name)

    temprp2 = ''
    if name.lower() == "crono":
        temprp2 = 'C'
    elif name.lower() == "marle":
        temprp2 = 'K'
    elif name.lower() == "lucca":
        temprp2 = 'G'
    elif name.lower() == "frog":
        temprp2 = 'O'
    elif name.lower() == "robo":
        temprp2 = 'S'
    elif name.lower() == "ayla":
        temprp2 = 'W'
    elif name.lower() == "magus":
        temprp2 = 'AA'
    else:
        # print('[Process:', localvar, ']', 'Error: Invalid Name')
        return 0

    temprp22 = colNameToNum(temprp2)
    playerdatavolatile = Path("Characters.xlsx")
    playerdatavolatileobj = openpyxl.load_workbook(playerdatavolatile, data_only=True)
    Characters = playerdatavolatileobj.active
    playerdata = [any] * 13
    for i in range(13):
        tempstatus = Characters[str(get_column_letter(temprp22)) + str(3 + i)].value
        if tempstatus == 'HP':
            playerdata[i] = Characters[str(get_column_letter(temprp22 + 2)) + str(3 + i)].value
        elif tempstatus == 'MP':
            playerdata[i] = Characters[str(get_column_letter(temprp22 + 2)) + str(3 + i)].value
        else:
            playerdata[i] = Characters[str(get_column_letter(temprp22 + 1)) + str(3 + i)].value
            if i > 0:
                playerdata[i] = int(playerdata[i])

    # print(playerdata)
    # print('[Process:', localvar, ']', 'Success')
    return playerdata


def ReadItemData(typew, name):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'a')
    print('[Process:', localvar, ']', 'Success')


def UseItem(name, target, player=None, mod=None):
    global globalvar
    global party
    localvar = globalvar
    globalvar = globalvar + 1

    playerdatavolatile = Path("Characters.xlsx")
    playerdatavolatileobj = openpyxl.load_workbook(playerdatavolatile, data_only=True)
    Characters = playerdatavolatileobj.active

    # print('[Process:', localvar, ']', 'UseItem', name, player, target)

    # if Inventory(item, check) == 1:

    temprp2 = None
    temprpp0 = None
    temprpp1 = None
    temprpp2 = None
    if target.lower() == "crono":
        temprp2 = 'C'
    elif target.lower() == "marle":
        temprp2 = 'K'
    elif target.lower() == "lucca":
        temprp2 = 'G'
    elif target.lower() == "frog":
        temprp2 = 'O'
    elif target.lower() == "robo":
        temprp2 = 'S'
    elif target.lower() == "ayla":
        temprp2 = 'W'
    elif target.lower() == "magus":
        temprp2 = 'AA'

    if party[0].lower() == "crono":
        temprpp0 = 'C'
    elif party[0].lower() == "marle":
        temprpp0 = 'K'
    elif party[0].lower() == "lucca":
        temprpp0 = 'G'
    elif party[0].lower() == "frog":
        temprpp0 = 'O'
    elif party[0].lower() == "robo":
        temprpp0 = 'S'
    elif party[0].lower() == "ayla":
        temprpp0 = 'W'
    elif party[0].lower() == "magus":
        temprpp0 = 'AA'
    if party[1].lower() == "crono":
        temprpp1 = 'C'
    elif party[1].lower() == "marle":
        temprpp1 = 'K'
    elif party[1].lower() == "lucca":
        temprpp1 = 'G'
    elif party[1].lower() == "frog":
        temprpp1 = 'O'
    elif party[1].lower() == "robo":
        temprpp1 = 'S'
    elif party[1].lower() == "ayla":
        temprpp1 = 'W'
    elif party[1].lower() == "magus":
        temprpp1 = 'AA'
    if party[2].lower() == "crono":
        temprpp2 = 'C'
    elif party[2].lower() == "marle":
        temprpp2 = 'K'
    elif party[2].lower() == "lucca":
        temprpp2 = 'G'
    elif party[2].lower() == "frog":
        temprpp2 = 'O'
    elif party[2].lower() == "robo":
        temprpp2 = 'S'
    elif party[2].lower() == "ayla":
        temprpp2 = 'W'
    elif party[2].lower() == "magus":
        temprpp2 = 'AA'

    if temprp2 is None:
        # print('[Process:', localvar, ']', 'Error: name does not match')
        return 0

    temprp22 = colNameToNum(temprp2)
    temprpp01 = colNameToNum(temprpp0)
    temprpp11 = colNameToNum(temprpp1)
    temprpp21 = colNameToNum(temprpp2)
    # print(temprpp0, temprpp1, temprpp2)

    bonus1 = 0
    status1 = ''
    bonus2 = 0
    status2 = ''
    bonus3 = 0
    status3 = ''
    targetype = ''

    # Preference of status: N1 Heal/Mp/Buffs/Tabs N2 Mp/Buffs N3/ Buffs

    if name.lower() == 'Tonic'.lower():
        # Heal 50 hp to one target
        bonus1 = 50
        status1 = 'HP'
        targetype = 'single'
    elif name.lower() == 'Mid Tonic'.lower():
        # Heal 200 hp to one target
        bonus1 = 200
        status1 = 'HP'
        targetype = 'single'
    elif name.lower() == 'Full Tonic'.lower():
        # Heal 500 hp to one target
        bonus1 = 500
        status1 = 'HP'
        targetype = 'single'
    elif name.lower() == 'Ether'.lower():
        # Heal 10 mp to one target
        bonus1 = 10
        status1 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Mid Ether'.lower():
        # Heal 30 mp to one target
        bonus1 = 30
        status1 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Full Ether'.lower():
        # Heal 60 mp to one target
        bonus1 = 60
        status1 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Hiper Ether'.lower():
        # Heal 100 mp to one target
        bonus1 = 100
        status1 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Elixir'.lower():
        # Heal 100 hp and 60 mp to one target
        bonus1 = 100
        status1 = 'HP'
        bonus2 = 60
        status2 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Mid Elixir'.lower():
        # Heal 500 hp and 100 mp to one target
        bonus1 = 500
        status1 = 'HP'
        bonus2 = 100
        status2 = 'MP'
        targetype = 'single'
    elif name.lower() == 'Mega Elixir'.lower():
        # Heal 500 hp and 100 mp to group
        bonus1 = 500
        status1 = 'HP'
        bonus2 = 100
        status2 = 'MP'
        targetype = 'group'
    elif name.lower() == 'Ambrosia'.lower():
        # Heal (1-500) hp and (1-100) mp to (1/2)single (1/2)group
        bonus1 = randint(1, 500)
        status1 = 'HP'
        bonus2 = randint(1, 100)
        status2 = 'MP'
        if randint(1, 2) == 2:
            targetype = 'group'
        else:
            targetype = 'single'
    elif name.lower() == 'Lapis'.lower():
        # Heal 200 hp to group
        bonus1 = 200
        status1 = 'HP'
        targetype = 'group'
    elif name.lower() == 'Shelter'.lower():
        # Heal 9999 hp and 9999 mp to group
        # Can only be used in safe zones
        bonus1 = 9999
        status1 = 'HP'
        bonus2 = 9999
        status2 = 'MP'
        targetype = 'group'
    elif name.lower() == 'Shield'.lower():
        # Buffs +0.33 def to one target
        bonus1 = 0.33
        status1 = 'PHYDAMRED'
        targetype = 'single'
    elif name.lower() == 'Barrier'.lower():
        # Buffs +0.33 mdef to one target
        bonus1 = 0.33
        status1 = 'MAGDAMRED'
        targetype = 'single'
    elif name.lower() == 'Revive'.lower():
        # Revive with 50 hp one target
        bonus1 = 50
        status1 = 'REVIVE'
        targetype = 'single'
    elif name.lower() == 'Heal'.lower():
        # Cure status one target
        bonus1 = 0
        status1 = 'STATUS'
        targetype = 'single'
    elif name.lower() == 'Magic Tab'.lower():
        # Buffs +10 MAG to one target for one battle
        bonus1 = 10
        status1 = 'MAG'
        targetype = 'single'
    elif name.lower() == 'Power Tab'.lower():
        # Buffs +10 PWR to one target for one battle
        bonus1 = 10
        status1 = 'PWR'
        targetype = 'single'
    elif name.lower() == 'Speed Tab'.lower():
        # Buffs +3 SPD to one target for one battle
        bonus1 = 3
        status1 = 'SPD'
        targetype = 'single'
    if mod.lower() == 'defense':
        bonus1 = 0.15
        status1 = 'PHYDAMRED'
        bonus2 = 0.15
        status2 = 'MAGDAMRED'
        targetype = 'single'

    if status1 == 'HP' or status2 == 'HP':
        if status2 == 'HP':
            bonus1 = bonus2
        tempheal = bonus1
        if targetype == 'single':
            temp = int(Characters[str(get_column_letter(temprp22 + 2)) + str(4)].value)
            if temp == 0:
                sumhp = 0
            else:
                sumhp = temp + bonus1
            maxhp = int(Characters[str(get_column_letter(temprp22 + 1)) + str(4)].value)
            if sumhp > maxhp:
                sumhp = maxhp
                tempheal = maxhp - temp
            # print('[Process:', localvar, ']', '+' + str(tempheal), target)
            Characters[str(get_column_letter(temprp22 + 2)) + str(4)] = sumhp
        elif targetype == 'group':
            # print('[Process:', localvar, ']', 'GroupHeal', party)
            temp = int(Characters[str(get_column_letter(temprpp01 + 2)) + str(4)].value)
            # print('debug', temp + bonus1)
            if temp == 0:
                sumhp = 0
            else:
                sumhp = temp + bonus1
            maxhp = int(Characters[str(get_column_letter(temprpp01 + 1)) + str(4)].value)
            if sumhp >= maxhp:
                sumhp = maxhp
                tempheal = maxhp - temp
            # print('[Process:', localvar, ']', '+' + str(tempheal), party[0])
            # print(get_column_letter(temprpp01 + 2), sumhp, Characters[str(get_column_letter(temprpp01 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp01 + 2)) + str(4)] = sumhp

            temp1 = int(Characters[str(get_column_letter(temprpp11 + 2)) + str(4)].value)
            # print('debug', temp + bonus1)
            if temp1 == 0:
                sumhp1 = 0
            else:
                sumhp1 = temp1 + bonus1
            maxhp1 = int(Characters[str(get_column_letter(temprpp11 + 1)) + str(4)].value)
            if sumhp1 >= maxhp1:
                sumhp1 = maxhp1
                tempheal = maxhp1 - temp1
            # print('[Process:', localvar, ']', '+' + str(tempheal), party[1])
            # print(get_column_letter(temprpp11 + 2), sumhp1, Characters[str(get_column_letter(temprpp11 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp11 + 2)) + str(4)] = sumhp1

            temp2 = int(Characters[str(get_column_letter(temprpp21 + 2)) + str(4)].value)
            # print('debug', temp + bonus1)
            if temp2 == 0:
                sumhp2 = 0
            else:
                sumhp2 = temp2 + bonus1
            maxhp2 = int(Characters[str(get_column_letter(temprpp21 + 1)) + str(4)].value)
            if sumhp2 >= maxhp2:
                sumhp2 = maxhp2
                tempheal = maxhp2 - temp2
            # print('[Process:', localvar, ']', '+' + str(tempheal), party[2])
            # print(get_column_letter(temprpp21 + 2), sumhp2, Characters[str(get_column_letter(temprpp21 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp21 + 2)) + str(4)] = sumhp2
        else:
            return 0
    if status1 == 'MP' or status2 == 'MP':
        if status2 == 'MP':
            bonus1 = bonus2
        tempmp = bonus1
        if targetype == 'single':
            temp = int(Characters[str(get_column_letter(temprp22 + 2)) + str(5)].value)
            maxmp = int(Characters[str(get_column_letter(temprp22 + 1)) + str(5)].value)
            summp = temp + bonus1
            if summp > maxmp:
                summp = maxmp
                tempmp = maxmp - temp
            # print('[Process:', localvar, ']', '+' + str(tempmp), target)
            Characters[str(get_column_letter(temprp22 + 2)) + str(5)] = summp
        elif targetype == 'group':
            # print('[Process:', localvar, ']', 'GroupHeal', party)
            temp = int(Characters[str(get_column_letter(temprpp01 + 2)) + str(5)].value)
            # print('debug', temp + bonus1)
            summp = temp + bonus1
            maxmp = int(Characters[str(get_column_letter(temprpp01 + 1)) + str(5)].value)
            if summp >= maxmp:
                summp = maxmp
                tempmp = maxmp - temp
            # print('[Process:', localvar, ']', '+' + str(tempmp), party[0])
            # print(get_column_letter(temprpp01 + 2), sumhp, Characters[str(get_column_letter(temprpp01 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp01 + 2)) + str(5)] = summp

            temp1 = int(Characters[str(get_column_letter(temprpp11 + 2)) + str(5)].value)
            # print('debug', temp + bonus1)
            summp1 = temp1 + bonus1
            maxmp1 = int(Characters[str(get_column_letter(temprpp11 + 1)) + str(5)].value)
            if summp1 >= maxmp1:
                summp1 = maxmp1
                tempmp = maxmp1 - temp1
            # print('[Process:', localvar, ']', '+' + str(tempmp), party[1])
            # print(get_column_letter(temprpp11 + 2), sumhp1, Characters[str(get_column_letter(temprpp11 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp11 + 2)) + str(5)] = summp1

            temp2 = int(Characters[str(get_column_letter(temprpp21 + 2)) + str(5)].value)
            # print('debug', temp + bonus1)
            summp2 = temp2 + bonus1
            maxmp2 = int(Characters[str(get_column_letter(temprpp21 + 1)) + str(5)].value)
            if summp2 >= maxmp2:
                sumhp2 = maxmp2
                tempheal = maxmp2 - temp2
            # print('[Process:', localvar, ']', '+' + str(tempmp), party[2])
            # print(get_column_letter(temprpp21 + 2), sumhp2, Characters[str(get_column_letter(temprpp21 + 2)) + str(4)].value)
            Characters[str(get_column_letter(temprpp21 + 2)) + str(5)] = summp2
        else:
            return 0
    if status1 == 'PHYDAMRED' or status2 == 'PHYDAMRED':
        if status2 == 'PHYDAMRED':
            bonus1 = bonus2
        temp = Characters[str(get_column_letter(temprp22 + 2)) + str(7)].value
        if temp is None:
            temp = 0
        temp = float(temp)
        sumdef = math.floor((temp + bonus1) * 100) / 100.0
        # print('sumdef1', sumdef)
        actualdef = int(Characters[str(get_column_letter(temprp22 + 1)) + str(7)].value)
        if 255 <= actualdef * (1 + sumdef):
            sumdef = math.floor(float((255 - actualdef) / actualdef) * 100) / 100.0
            # print('sumdef2', sumdef)
        # print('[Process:', localvar, ']', '+' + str(sumdef), target)
        Characters[str(get_column_letter(temprp22 + 2)) + str(7)] = sumdef
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'MAGDAMRED' or status2 == 'MAGDAMRED':
        if status2 == 'MAGDAMRED':
            bonus1 = bonus2
        temp = Characters[str(get_column_letter(temprp22 + 2)) + str(15)].value
        if temp is None:
            temp = 0
        temp = float(temp)
        sumdef = math.floor((temp + bonus1) * 100) / 100.0
        # print('sumdef1', sumdef)
        actualdef = int(Characters[str(get_column_letter(temprp22 + 1)) + str(15)].value)
        if 255 <= actualdef * (1 + sumdef):
            sumdef = math.floor(float((255 - actualdef) / actualdef) * 100) / 100.0
            # print('sumdef2', sumdef)
        # print('[Process:', localvar, ']', '+' + str(sumdef), target)
        Characters[str(get_column_letter(temprp22 + 2)) + str(15)] = sumdef
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'REVIVE':
        temp = int(Characters[str(get_column_letter(temprp22 + 2)) + str(4)].value)
        temp2 = int(Characters[str(get_column_letter(temprp22 + 1)) + str(4)].value)
        if temp2 <= bonus1:
            bonus1 = temp2
        Characters[str(get_column_letter(temprp22 + 2)) + str(4)] = temp + bonus1
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'STATUS':
        temp = Characters[str(get_column_letter(temprp22 + 1)) + str(14)].value
        print('debug', temp + bonus1)
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'MAG':
        temp = Characters[str(get_column_letter(temprp22 + 2)) + str(14)].value
        temp2 = int(Characters[str(get_column_letter(temprp22 + 1)) + str(14)].value)
        if temp is None:
            temp = 0
        temp = int(temp)
        sumMAG = temp + bonus1
        if sumMAG + temp2 > 99:
            sumMAG = temp2 - 99
        # print('sumdef1', sumdef)
        # print('[Process:', localvar, ']', '+' + str(sumMAG), target)
        Characters[str(get_column_letter(temprp22 + 2)) + str(14)] = sumMAG
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'PWR':
        temp = Characters[str(get_column_letter(temprp22 + 2)) + str(9)].value
        temp2 = int(Characters[str(get_column_letter(temprp22 + 1)) + str(9)].value)
        if temp is None:
            temp = 0
        temp = int(temp)
        sumPWR = temp + bonus1
        if sumPWR + temp2 > 99:
            sumPWR = temp2 - 99
        # print('sumdef1', sumdef)
        # print('[Process:', localvar, ']', '+' + str(sumPWR), target)
        Characters[str(get_column_letter(temprp22 + 2)) + str(9)] = sumPWR
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    if status1 == 'SPD':
        temp = Characters[str(get_column_letter(temprp22 + 2)) + str(10)].value
        temp2 = int(Characters[str(get_column_letter(temprp22 + 1)) + str(10)].value)
        if temp is None:
            temp = 0
        temp = int(temp)
        sumSPD = temp + bonus1
        if sumSPD + temp2 > 99:
            sumSPD = temp2 - 99
        # print('sumdef1', sumdef)
        # print('[Process:', localvar, ']', '+' + str(sumSPD), target)
        Characters[str(get_column_letter(temprp22 + 2)) + str(10)] = sumSPD
        # Characters[str(get_column_letter(temprp22+1)) + str(14)] = temp + bonus1
    # Inventory(name, remove)

    playerdatavolatileobj.save(playerdatavolatile)
    # print('[Process:', localvar, ']', 'Success')


def AtkFormula(name, Power, Level, Hit, Weapon):
    if name.lower() == 'crono' or name.lower() == 'frog' or name.lower() == 'magus' or name.lower() == 'robo':
        # print("man")
        return int((Power * 4 / 3) + (Weapon * 5 / 9))
    elif name.lower() == 'lucca' or name.lower() == 'marle':
        return int((Hit + Weapon) * 2 / 3)
    elif name.lower() == 'ayla':
        return int((Power * 1.75) + (Level * Level / 45.5))


def EquipItem(typew, name, player):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'EquipItem', typew, name, player)

    Weapons = Path("Weapons.xlsx")
    Weaponsobj = openpyxl.load_workbook(Weapons, data_only=True)
    Wpn = Weaponsobj.active

    Items = Path("Items.xlsx")
    Itemsobj = openpyxl.load_workbook(Items, data_only=True)
    Itm = Itemsobj.active

    playerdatavolatile = Path("Characters.xlsx")
    playerdatavolatileobj = openpyxl.load_workbook(playerdatavolatile, data_only=True)
    Characters = playerdatavolatileobj.active

    temprp1 = ''
    temprp2 = ''
    if player.lower() == "crono":
        temprp1 = 'C'
        temprp2 = 'C'
    elif player.lower() == "marle":
        temprp1 = 'J'
        temprp2 = 'K'
    elif player.lower() == "lucca":
        temprp1 = 'Q'
        temprp2 = 'G'
    elif player.lower() == "frog":
        temprp1 = 'X'
        temprp2 = 'O'
    elif player.lower() == "robo":
        temprp1 = 'AE'
        temprp2 = 'S'
    elif player.lower() == "ayla":
        temprp1 = 'AL'
        temprp2 = 'W'
    elif player.lower() == "magus":
        temprp1 = 'AR'
        temprp2 = 'AA'
    temprp12 = int(colNameToNum(temprp1))
    temprp21 = int(colNameToNum(temprp2))

    if typew.lower() == 'weapon':
        for i in range(1024):
            count = str(int(i + 4))
            const = Wpn[get_column_letter(temprp12 + 1) + count].value
            if const is None:
                break
            # print(get_column_letter(temprp12 + 1), count)
            if const.lower() == name.lower():
                atkline = str(6)
                # print('debug', get_column_letter(temprp21 + 1))
                tempa = int(Characters[str(get_column_letter(temprp21 + 1)) + '9'].value)
                tempb = int(Characters[str(get_column_letter(temprp21 + 1)) + '8'].value)
                tempc = int(Characters[str(get_column_letter(temprp21 + 1)) + '11'].value)
                if player.lower() == 'ayla':
                    tempd = 0
                else:
                    tempd = int(Wpn[get_column_letter(temprp12 + 4) + count].value)

                # print('debug2', name, tempa, tempb, tempc, tempd, 'debug2', AtkFormula(name, tempa, tempb, tempc, tempd))
                Characters[str(get_column_letter(temprp21 + 1)) + atkline] = AtkFormula(player, tempa, tempb, tempc,
                                                                                        tempd)
                Characters[str(get_column_letter(temprp21 + 1)) + '18'] = name
                Characters[str(get_column_letter(temprp21 + 2)) + '18'] = tempd

                print('[Process:', localvar, ']', 'Success')
                break
        # print(str(Wpn[get_column_letter(temprp12 + 1) + count].value))
        # print(get_column_letter(temprp12 + 1))
    elif typew.lower() == 'helm' or typew.lower() == 'armor' or typew.lower() == 'accessory':
        for i in range(1024):
            count = str(i + 4)
            const = None
            const2 = None
            if typew.lower() == 'helm':
                temprp12 = int(colNameToNum('C'))
                const = Itm[get_column_letter(temprp12 + 1) + count].value
                const2 = '19'
            elif typew.lower() == 'armor':
                temprp12 = int(colNameToNum('K'))
                const = Itm[get_column_letter(temprp12 + 1) + count].value
                const2 = '20'
            elif typew.lower() == 'accessory':
                temprp12 = int(colNameToNum('S'))
                const = Itm[get_column_letter(temprp12 + 1) + count].value
                const2 = '21'
            if const is None:
                print('[Process:', localvar, ']', 'Error: typew.lower() does not match')
                break
            if const == name:
                # print('debug', const, const2, name)
                Characters[str(get_column_letter(temprp21 + 1)) + const2] = name
                if typew.lower() != 'accessory':
                    Characters[str(get_column_letter(temprp21 + 2)) + const2] = Itm[
                        get_column_letter(temprp12 + 4) + count].value
                    defhelm = Characters[str(get_column_letter(temprp21 + 2)) + '19'].value
                    defarmor = Characters[str(get_column_letter(temprp21 + 2)) + '20'].value
                    if defhelm is None:
                        defhelm = 0
                        Characters[str(get_column_letter(temprp21 + 2)) + '19'] = 0
                    if defarmor is None:
                        defarmor = 0
                        Characters[str(get_column_letter(temprp21 + 2)) + '20'] = 0
                    Characters[str(get_column_letter(temprp21 + 1)) + '7'] = defarmor + defhelm
                print('[Process:', localvar, ']', 'Success')
                break
            # print(get_column_letter(temprp12 + 1), count)
        # print(str(Wpn[get_column_letter(temprp12 + 1) + count].value))
        # print(get_column_letter(temprp12 + 1))

    playerdatavolatileobj.save(playerdatavolatile)


def LevelUpPlayer(name, level=1, mod='', amount=0):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'UpdatePlayer', name)

    temprp1 = str
    temprp2 = str

    if name.lower() == "crono":
        temprp1 = 'B'
        temprp2 = 'C'
    elif name.lower() == "marle":
        temprp1 = 'AP'
        temprp2 = 'K'
    elif name.lower() == "lucca":
        temprp1 = 'V'
        temprp2 = 'G'
    elif name.lower() == "frog":
        temprp1 = 'L'
        temprp2 = 'O'
    elif name.lower() == "robo":
        temprp1 = 'AZ'
        temprp2 = 'S'
    elif name.lower() == "ayla":
        temprp1 = 'BJ'
        temprp2 = 'W'
    elif name.lower() == "magus":
        temprp1 = 'AF'
        temprp2 = 'AA'
    else:
        print('[Process:', localvar, ']', 'Error: Invalid Name')
        return 0

    playerdata = Path("Player_Status_Techs.xlsx")
    playerdataobj = openpyxl.load_workbook(playerdata, data_only=True)
    Player_Status_Techs = playerdataobj.active

    playerdatavolatile = Path("Characters.xlsx")
    playerdatavolatileobj = openpyxl.load_workbook(playerdatavolatile, data_only=True)
    Characters = playerdatavolatileobj.active

    # print(temprp1)

    temprp12 = colNameToNum(temprp1)  # Reading columns values in Excel is hard
    temprp22 = colNameToNum(temprp2)

    if mod == 'add':
        Characters[str(get_column_letter(temprp22 + 1)) + '16'] = Characters[str(get_column_letter(temprp22 + 1)) + '16'] + amount
    # print("HP: " + str(Player_Status_Techs[str(get_column_letter(temprp12 + 1)) + "5"].value))

    leveline = str(4 + level)
    for i in range(0, 255):
        if type(Player_Status_Techs[str(get_column_letter(temprp12 + i)) + leveline].value) != int:
            break
        const = str(Player_Status_Techs[str(get_column_letter(temprp12 + i)) + "4"].value)
        const2 = str(
            Player_Status_Techs[str(get_column_letter(temprp12 + i)) + leveline].value)
        # print(i, const, const2) # The Debug Master Print
        # Characters[str(get_column_letter(temprp22+1)) + str(3+i)] = Player_Status_Techs[str(get_column_letter(temprp12 + i)) + "5"].value

        temprp3 = str(Characters[str(get_column_letter(temprp22)) + str(4 + i)].value)
        temprp4 = ''
        if temprp3 == "Weapon" or temprp3 == "Helmet" or temprp3 == "Armor" or temprp3 == "Accessory":
            break
        if const == "HP":
            temprp4 = str(4)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = const2
        elif const == "MP":
            temprp4 = str(5)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = const2
        elif const == "Evasion":
            temprp4 = str(12)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Hit":
            temprp4 = str(11)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Magic":
            temprp4 = str(14)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Magic Defense":
            temprp4 = str(15)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Power":
            temprp4 = str(9)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Stamina":
            temprp4 = str(13)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        elif const == "Level":
            temprp4 = str(8)
            Characters[str(get_column_letter(temprp22 + 2)) + temprp4] = 0
        Characters[str(get_column_letter(temprp22 + 2)) + '7'] = 0
        Characters[str(get_column_letter(temprp22 + 2)) + '10'] = 0
        Characters[str(get_column_letter(temprp22 + 1)) + '16'] = 0
        Characters[str(get_column_letter(temprp22 + 1)) + '17'] = 0
        # print('temprp4', temprp4)
        # print('Characters', str(Characters[str(get_column_letter(temprp22)) + temprp4].value))
        Characters[str(get_column_letter(temprp22 + 1)) + temprp4] = const2
        # print("\n")

    atkline = str(6)
    # print('debug', get_column_letter(temprp21 + 1))
    tempa = int(Characters[str(get_column_letter(temprp22 + 1)) + '9'].value)
    tempb = int(Characters[str(get_column_letter(temprp22 + 1)) + '8'].value)
    tempc = int(Characters[str(get_column_letter(temprp22 + 1)) + '11'].value)
    tempd = int(Characters[str(get_column_letter(temprp22 + 2)) + '18'].value)

    # print('debug1', get_column_letter((temprp22+1)))
    # print('debug1', name, tempa, tempb, tempc, tempd, 'debug1', AtkFormula(name, tempa, tempb, tempc, tempd))
    Characters[str(get_column_letter(temprp22 + 1)) + atkline] = AtkFormula(name, tempa, tempb, tempc, tempd)

    print('[Process:', localvar, ']', 'Success')
    playerdataobj.save(playerdata)
    playerdatavolatileobj.save(playerdatavolatile)


def battle(maps, mob1=None, mob2=None, mob3=None, mob4=None, mob5=None, mod=None):
    global globalvar
    localvar = globalvar
    globalvar = globalvar + 1
    print('[Process:', localvar, ']', 'Battle', maps, mob1, mob2, mob3, mob4, mob5)

    playerdatavolatile = Path("Characters.xlsx")
    playerdatavolatileobj = openpyxl.load_workbook(playerdatavolatile, data_only=True)
    Characters = playerdatavolatileobj.active
    temprpp0 = ''
    temprpp1 = ''
    temprpp2 = ''
    techp = ['', '', '']
    if party[0].lower() == "crono":
        temprpp0 = 'C'  # Related to player data
        techp[0] = 'K'  # Related to TP
    elif party[0].lower() == "marle":
        temprpp0 = 'K'
        techp[0] = 'M'
    elif party[0].lower() == "lucca":
        temprpp0 = 'G'
        techp[0] = 'L'
    elif party[0].lower() == "frog":
        temprpp0 = 'O'
        techp[0] = 'N'
    elif party[0].lower() == "robo":
        temprpp0 = 'S'
        techp[0] = 'O'
    elif party[0].lower() == "ayla":
        temprpp0 = 'W'
        techp[0] = 'P'
    elif party[0].lower() == "magus":
        temprpp0 = 'AA'
        techp[0] = 'Q'
    if party[1].lower() == "crono":
        temprpp1 = 'C'
        techp[1] = 'K'
    elif party[1].lower() == "marle":
        temprpp1 = 'K'
        techp[1] = 'M'
    elif party[1].lower() == "lucca":
        temprpp1 = 'G'
        techp[1] = 'L'
    elif party[1].lower() == "frog":
        temprpp1 = 'O'
        techp[1] = 'N'
    elif party[1].lower() == "robo":
        temprpp1 = 'S'
        techp[1] = 'O'
    elif party[1].lower() == "ayla":
        temprpp1 = 'W'
        techp[1] = 'P'
    elif party[1].lower() == "magus":
        temprpp1 = 'AA'
        techp[1] = 'Q'
    if party[2].lower() == "crono":
        temprpp2 = 'C'
        techp[2] = 'K'
    elif party[2].lower() == "marle":
        temprpp2 = 'K'
        techp[2] = 'M'
    elif party[2].lower() == "lucca":
        temprpp2 = 'G'
        techp[2] = 'L'
    elif party[2].lower() == "frog":
        temprpp2 = 'O'
        techp[2] = 'N'
    elif party[2].lower() == "robo":
        temprpp2 = 'S'
        techp[2] = 'O'
    elif party[2].lower() == "ayla":
        temprpp2 = 'W'
        techp[2] = 'P'
    elif party[2].lower() == "magus":
        temprpp2 = 'AA'
        techp[2] = 'Q'
    temprpp01 = colNameToNum(temprpp0)
    temprpp11 = colNameToNum(temprpp1)
    temprpp21 = colNameToNum(temprpp2)
    # Uses a dinamical file to register buffs/mobs hp and etc
    monster = list
    mobn = int
    if mob1 is None:
        if maps.lower() == 'guardia florest - present':
            print('5d20, +1 monster for each 5+,5+,10+,15+,20 cap at 5 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd20, each value translate into one different mob')
            mob = [1] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 9:
                    monster[i] = 'Amanita'
                elif 9 <= mob[i] < 16:
                    monster[i] = 'Scarab'
                else:
                    monster[i] = 'Gilded Bellbird'
            # print('debug', mobn, monster)
        if maps.lower() == 'leene square - present':
            confirm = input('Do you wanna fight Gato(Y/N)? ')
            if confirm.lower() == 'y':
                mobn = 1
                monster = [''] * mobn
                monster[0] = 'Gato'
                # print(str(mobn) + 'd20, each value translate into one different mob')
                # print('debug', mobn, monster)
            else:
                return 0
        if maps.lower() == 'prison towers - present':
            print('5d20, +1 monster for each 5+,5+,15+,15+,20 cap at 5 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd20, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 11:
                    monster[i] = 'Royal Guard'
                elif 11 <= mob[i] < 14:
                    monster[i] = 'Sentry'
                elif 14 <= mob[i] < 20:
                    monster[i] = 'Lancer'
                else:
                    monster[i] = 'Gaoler'
            # print('debug', mobn, monster)
        if maps.lower() == 'heckran cave - present':
            print('5d20, +1 monster for each 5+,7+,10+,15+,20 cap at 5 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd20, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 5:
                    monster[i] = 'Cave Stalker'
                elif 5 <= mob[i] < 9:
                    monster[i] = 'Djinn Bottle'
                elif 9 <= mob[i] < 15:
                    monster[i] = 'Cave Bat'
                elif 15 <= mob[i] < 19:
                    monster[i] = 'Boundillo'
                else:
                    monster[i] = 'Rhino Weevil'
            # print('debug', mobn, monster)
        if maps.lower() == 'northern ruins - present':
            print('3d20, +1 monster for each 5+,5+,10+ cap at 3 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd100, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 99:
                    monster[i] = 'Royal Guard'
                else:
                    monster[i] = 'Cyrus'
            # print('debug', mobn, monster)
        if maps.lower() == 'truce Canyon - middle age':
            print('6d20, +1 monster for each 5+,5+,5+,5+,5+ cap at 6 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd100, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 30:
                    monster[i] = 'Blue Imp'
                if 30 <= mob[i] < 50:
                    monster[i] = 'Green Imp'
                if 50 <= mob[i] < 70:
                    monster[i] = 'Roundillo'
                if 70 <= mob[i] < 90:
                    monster[i] = 'Roundillo Rider'
                else:
                    monster[i] = 'Fiendillo'
            # print('debug', mobn, monster)
        if maps.lower() == 'guardia florest - middle age':
            print('5d20, +1 monster for each 5+,5+,5+,10+,15+ cap at 5 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd100, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 40:
                    monster[i] = 'Green Imp'
                elif 40 <= mob[i] < 60:
                    monster[i] = 'Roundillo'
                elif 60 <= mob[i] < 70:
                    monster[i] = 'Roundillo Rider'
                elif 70 <= mob[i] < 80:
                    monster[i] = 'Blue Eaglet'
                else:
                    monster[i] = 'Naga'
            # print('debug', mobn, monster)
        if maps.lower() == 'cathedral - middle age':
            print('6d20, +1 monster for each 5+,5+,5+,10+,10+,15+ cap at 6 monsters')
            mobn = int(input('How many monsters? '))
            print(str(mobn) + 'd100, each value translate into one different mob')
            mob = [0] * mobn
            monster = [''] * mobn
            for i in range(mobn):
                mob[i] = int(input('#' + str(i + 1) + ' mob: '))
                if 1 <= mob[i] < 30:
                    monster[i] = 'Naga'
                if 30 <= mob[i] < 45:
                    monster[i] = 'Viper'
                if 45 <= mob[i] < 70:
                    monster[i] = 'Diablo'
                if 70 <= mob[i] < 80:
                    monster[i] = 'Underling (Blue)'
                else:
                    monster[i] = 'Mad Bat'
            # print('debug', mobn, monster)
        if maps.lower() == 'yakra - middle age':
            confirm = input('Do you wanna fight Yakra(Y/N)? ')
            if confirm.lower() == 'y':
                mobn = 1
                monster = [''] * mobn
                monster[0] = 'Yakra'
                # print(str(mobn) + 'd20, each value translate into one different mob')
                # print('debug', mobn, monster)
            else:
                return 0
            # print('debug', mobn, monster)

    monsterdata = [''] * len(monster)
    ActionTimer = [''] * (len(monster) + len(party))
    ActionTimer[0] = Characters[get_column_letter(temprpp01 + 1) + str(10)].value
    ActionTimer[1] = Characters[get_column_letter(temprpp11 + 1) + str(10)].value
    ActionTimer[2] = Characters[get_column_letter(temprpp21 + 1) + str(10)].value
    for j in range(len(monster)):
        if monster[j] == 'Yakra':
            monsterdata[j] = ReadBossData(monster[j])
        else:
            monsterdata[j] = ReadMonsterData(monster[j])
        ActionTimer[j + 3] = int(str(monsterdata[j][11]).replace('[', '').replace(']', ''))
    for j in range(len(ActionTimer)):
        ActionTimer[j] = maxspeed - int(ActionTimer[j])  # Max Speed is 20
    # print(ActionTimer)
    k = 2
    playturn = []
    while k < 512:
        for l in range(len(ActionTimer)):
            # print(k, l, ActionTimer[l])
            if l <= 2 and k % int(ActionTimer[l]) == 0:
                # print('valid')
                playturn.append(str(party[l]))
            elif k % int(ActionTimer[l]) == 0:
                # print('valid', l, ActionTimer[l])
                playturn.append(int(l - 3))
        k = k + 1
    # print('playturn', len(playturn), playturn)   # I love this variable
    # print('debug', monsterdata[0])
    lose = False
    for i in range(255):
        waitvariable = input('')
        os.system('cls')
        losecounter = 0
        victorycounter = 0
        for j in range(len(party)):
            templayer = ReadPlayerData(str(party[j]))
            if int(templayer[1]) <= 0:
                losecounter = losecounter + 1
        if losecounter == 3 or lose is True:
            print('All your heroes died')
            wait = input()
            break
        for j in range(len(monster)):
            if monster[j] == 'dead' or monster[j] == 'flew':
                victorycounter = victorycounter + 1
        # print('victorycounter', victorycounter, monster)
        if victorycounter == len(monster):
            # PLACE LOOT TABLE THERE FOR MY SANITY AND IM GOING TO BATH
            print('VICTORY')
            lootbag = []
            totalxp = 0
            totaltp = 0
            totalgold = 0
            for j in range(len(monster)):
                if monster[j] == 'dead':
                    # print(monsterdata)
                    # print(monsterdata[j])
                    # print(monsterdata[j][5])
                    for k in range(len(monsterdata[j][5])):
                        if str(monsterdata[j][5][k]) != 'None':
                            lootbag.append(str(monsterdata[j][5][k]))
                    totalxp = totalxp + int(monsterdata[j][1][0])
                    totaltp = totaltp + int(monsterdata[j][2][0])
                    totalgold = totalgold + int(monsterdata[j][3][0])

            # print(totalxp, totaltp, totalgold, lootbag)

            # SAVING THE XP GAIN
            # print(get_column_letter(temprpp01+1), Characters[get_column_letter(temprpp01 + 1) + '16'].value)
            Characters[get_column_letter(temprpp01 + 1) + '16'] = int(Characters[get_column_letter(
                temprpp01 + 1) + '16'].value) + totalxp
            Characters[get_column_letter(temprpp11 + 1) + '16'] = int(Characters[get_column_letter(
                temprpp11 + 1) + '16'].value) + totalxp
            Characters[get_column_letter(temprpp21 + 1) + '16'] = int(Characters[get_column_letter(
                temprpp21 + 1) + '16'].value) + totalxp

            # PRINTING THE LOOT TABLE
            print('\nLOOT:')
            for j in range(len(lootbag)):
                print('[' + str(j + 1) + ']', lootbag[j])
            print('XP: ' + str(totalxp))
            print('TP: ' + str(totaltp))
            print('GOLD: ' + str(totalgold))
            for j in range(len(lootbag)):
                Inventory('add', lootbag[j])
            Inventory('add', int(totalgold), 'gold')

            # SAVING THE TECH POINTS GAIN
            TechListabc = Path("TechList.xlsx")
            TechListabcobj = openpyxl.load_workbook(TechListabc, data_only=True)
            TechList = TechListabcobj.active
            # print(techp)
            for j in range(3):
                TechList[techp[j] + '4'] = int(TechList[techp[j] + '4'].value) + totaltp
            playerdatavolatileobj.save(playerdatavolatile)
            TechListabcobj.save(TechListabc)

            wait = input()
            break

        print('\n[Round'.upper() + str(i + 1).upper() + ']'.upper() + '\nEnemies:')
        for j in range(len(monster)):
            showhp = ''
            # if monster[j] == 'dead':
            #     print('[' + str(j + 1) + '] ' + 'dead')
            # if monster[j] == 'flew':
            #     print('[' + str(j + 1) + '] ' + 'flew')
            # else:
            if 'vision' in mod:
                showhp = ' HP: ' + str(monsterdata[j][0]).replace('[', '').replace(']', '')
            print('[' + str(j + 1) + '] ' + str(monster[j]) + showhp)
        print('Party:')
        # print(party)
        # print(ReadPlayerData('crono'))
        for j in range(len(party)):
            templayer = ReadPlayerData(str(party[j]))
            print('[' + str(j + 1) + '] ' + str(party[j]) + ' HP: ' + str(templayer[1]) + ' MP: ' + str(templayer[2]))

        while 1 != 0:
            if type(playturn[i]) == str and playturn[i].lower() in party:
                templayer = ReadPlayerData(str(playturn[i]))
                print('\n' + playturn[i].upper(), 'turn')
                if int(templayer[1]) <= 0:
                    print(playturn[i].upper(), 'is dead')
                    break
                print('[1] Attack')
                print('[2] Techs')
                print('[3] Items')
                print('[4] Defend')
                chooseaction = input()

                if chooseaction == '1':
                    print('Escolha um alvo:')
                    tempk = 0
                    for k in range(len(monster)):
                        tempk = k
                        showhp = ''
                        # if monster[k] == 'dead':
                        #     print('[' + str(k + 1) + '] ' + 'dead')
                        # if monster[k] == 'flew':
                        #     print('[' + str(k + 1) + '] ' + 'flew')
                        # else:
                        if 'vision' in mod:
                            showhp = ' HP: ' + str(monsterdata[k][0]).replace('[', '').replace(']', '')
                        print('[' + str(k + 1) + '] ' + str(monster[k]) + showhp)
                    print('[' + str(tempk + 2) + '] ' + 'Cancel')
                    chooseattack = int(input()) - 1
                    if chooseattack > tempk:
                        print('')
                    else:
                        # print(chooseattack, monsterdata[chooseattack][0][0], monsterdata)
                        if int(monsterdata[chooseattack][0][0]) > 0:
                            templife = int(monsterdata[chooseattack][0][0])
                            if playturn[i] == party[0]:
                                # print(monsterdata[chooseattack][8][0])
                                damage = templife - int(
                                    int(Characters[get_column_letter(temprpp01 + 1) + '6'].value) * (
                                            256 - int(monsterdata[chooseattack][8][0])) / 256)
                            elif playturn[i] == party[1]:
                                damage = templife - int(
                                    int(Characters[get_column_letter(temprpp11 + 1) + '6'].value) * (
                                            256 - int(monsterdata[chooseattack][8][0])) / 256)
                            elif playturn[i] == party[2]:
                                damage = templife - int(
                                    int(Characters[get_column_letter(temprpp21 + 1) + '6'].value) * (
                                            256 - int(monsterdata[chooseattack][8][0])) / 256)
                            else:
                                damage = 0
                            monsterdata[chooseattack][0][0] = damage
                            if int(monsterdata[chooseattack][0][0]) <= 0:
                                monsterdata[chooseattack][0][0] = 0
                                monster[chooseattack] = 'dead'
                            print(playturn[i], 'dealt', str((templife - damage) * -1), 'to the foe')
                            break
                        else:
                            print('Cannot attack that one')
                if chooseaction == '2':
                    # Using a tech
                    # print('some tech')
                    TechListabc = Path("TechList.xlsx")
                    TechListabcobj = openpyxl.load_workbook(TechListabc, data_only=True)
                    TechList = TechListabcobj.active
                    UnableTechs = []
                    datacolumn = ''

                    if playturn[i].lower() == 'crono':
                        datacolumn = 'C'
                    elif playturn[i].lower() == 'lucca':
                        datacolumn = 'D'
                    elif playturn[i].lower() == 'marle':
                        datacolumn = 'E'
                    elif playturn[i].lower() == 'frog':
                        datacolumn = 'F'
                    elif playturn[i].lower() == 'robo':
                        datacolumn = 'G'
                    elif playturn[i].lower() == 'ayla':
                        datacolumn = 'H'
                    elif playturn[i].lower() == 'magus':
                        datacolumn = 'I'
                    else:
                        print('Please learn how to write zzz')

                    k = 0
                    while TechList[datacolumn + str(4 + k)].value is not None:
                        UnableTechs.append(TechList[datacolumn + str(4 + k)].value)
                        k = k + 1
                    k = 0  # Be careful with while
                    # print('UnableTechs', UnableTechs)
                    print('Choose an tech to use')
                    tempabc = 0
                    for j in range(len(UnableTechs)):
                        tempabc = j
                        print('[' + str(j + 1) + '] ' + str(UnableTechs[j]) + ' MP COST:' + str(
                            tech(UnableTechs[j], mod='cost')))
                    print('[' + str(j + 2) + '] ' + 'Cancel')
                    choosetech = int(int(input()) - 1)
                    print('debug', choosetech, tempabc+1)
                    if choosetech >= tempabc+1:
                        print('')
                        # print('')
                    else:
                        # print('Choosedone', UnableTechs[choosetech])
                        if int(templayer[2]) < int(tech(UnableTechs[choosetech], mod='cost')):
                            print('Não possui mana o suficiente')
                            continue
                        # print(templayer)

                        print('Escolha um alvo:')
                        tempk = 0
                        for k in range(len(monster)):
                            tempk = k
                            showhp = ''
                            # if monster[k] == 'dead':
                            #     print('[' + str(k + 1) + '] ' + 'dead')
                            # if monster[k] == 'flew':
                            #     print('[' + str(k + 1) + '] ' + 'flew')
                            # else:
                            if 'vision' in mod:
                                showhp = ' HP: ' + str(monsterdata[k][0]).replace('[', '').replace(']', '')
                            print('[' + str(k + 1) + '] ' + str(monster[k]) + showhp)
                        print('[' + str(tempk + 2) + '] ' + 'Cancel')
                        chooseattack = int(input()) - 1
                        if chooseattack > tempk:
                            print('')
                            # print('')
                        else:
                            # print(chooseattack, monsterdata[chooseattack][0][0], monsterdata)
                            techdata = tech(UnableTechs[choosetech], HP=templayer[1], MP=templayer[2], ATK=templayer[3],
                                            DFN=templayer[4], LVL=templayer[5], PWR=templayer[6], SPD=templayer[7],
                                            HIT=templayer[8], EVA=templayer[9], STM=templayer[10], MAG=templayer[11],
                                            MDF=templayer[12], RNG=1)
                            techdamage = 0
                            templife = int(monsterdata[chooseattack][0][0])

                            if grouptechs(UnableTechs[choosetech]):
                                # print(monster)
                                confirmattack = False
                                for j in range(len(monster)):
                                    templife = int(monsterdata[j][0][0])
                                    if techdata[1] == 'PHY':
                                        techdamage = int(techdata[0] * (
                                                256 - int(monsterdata[j][8][0])) / 256)
                                    elif techdata[1] == 'LIG' or techdata[1] == 'SHA' or techdata[1] == 'FIR' or \
                                            techdata[
                                                1] == 'WAT':
                                        techdamage = int(techdata[0] * (
                                                101 - int(monsterdata[j][10][0])) / 101)
                                    if int(monsterdata[j][0][0]) > 0:
                                        damage = templife - techdamage
                                    else:
                                        damage = 0

                                    # print('debug', monsterdata[j][0][0], templife, techdamage, techdata[0], damage)
                                    monsterdata[j][0][0] = damage
                                    if int(monsterdata[j][0][0]) <= 0:
                                        monsterdata[j][0][0] = 0
                                        # print(type(i), type(j), type(choosetech), type(playturn[i]))
                                        print(playturn[i], 'using', UnableTechs[choosetech], 'did',
                                              str((templife - damage) * -1),
                                              'to', monster[j])
                                        monster[j] = 'dead'
                                        confirmattack = True
                                    else:
                                        print(playturn[i], 'using', UnableTechs[choosetech], 'did',
                                              str((templife - damage) * -1),
                                              'to', monster[j])
                                        confirmattack = True
                                if confirmattack is True:
                                    break

                            if techdata[1] == 'PHY':
                                techdamage = int(techdata[0] * (
                                        256 - int(monsterdata[chooseattack][8][0])) / 256)
                            elif techdata[1] == 'LIG' or techdata[1] == 'SHA' or techdata[1] == 'FIR' or techdata[
                                1] == 'WAT':
                                techdamage = int(techdata[0] * (
                                        101 - int(monsterdata[chooseattack][10][0])) / 101)
                            if int(monsterdata[chooseattack][0][0]) > 0:
                                damage = templife - techdamage
                            else:
                                damage = 0

                            monsterdata[chooseattack][0][0] = damage
                            # print('debug', monsterdata[chooseattack][0][0], templife, techdamage, techdata[0])
                            if int(monsterdata[chooseattack][0][0]) <= 0:
                                monsterdata[chooseattack][0][0] = 0
                                print(playturn[i], 'using', UnableTechs[choosetech], 'did',
                                      str((templife - damage) * -1),
                                      'to', monster[chooseattack])
                                monster[chooseattack] = 'dead'
                                break
                            else:
                                print(playturn[i], 'using', UnableTechs[choosetech], 'did',
                                      str((templife - damage) * -1),
                                      'to', monster[chooseattack])
                                break
                            TechListabcobj.save(TechListabc)

                if chooseaction == '3':
                    # Using a item
                    invlist = Inventory('list')
                    tempj = 0
                    for j in range(len(invlist)):
                        tempj = j
                        print('[' + str(j + 1) + '] ' + str(invlist[j][1]) + ' ' + str(invlist[j][2]))
                        if invlist[j + 1][1] == '':
                            break
                    print('[' + str(tempj + 2) + '] ' + 'Cancel')
                    chooseitem = int(input())
                    itemname = invlist[chooseitem - 1][2]
                    # print(chooseitem, tempj + 2, itemname)
                    if 0 < chooseitem < tempj + 2:
                        if Inventory('check', itemname) is True:
                            print('Whats your target?')
                            for j in range(len(party)):
                                print('[' + str(j + 1) + '] ' + party[j])
                            itemtarget = int(input()) - 1
                            if 0 <= itemtarget < 3:
                                templayer = ReadPlayerData(str(party[itemtarget]))
                                # print(templayer, itemname, party[itemtarget])
                                if itemname.lower() == 'revive' and int(templayer[1]) == 0:
                                    Inventory('remove', itemname)
                                    UseItem(itemname, target=party[itemtarget])
                                    break
                                elif itemname.lower() != 'revive':
                                    Inventory('remove', itemname)
                                    UseItem(itemname, target=party[itemtarget])
                                    break
                        else:
                            print('You dont have this item')
                if chooseaction == '4':
                    UseItem('', playturn[i], mod='defense')
                    print('Defense buff')
                    break
            else:
                if monster[playturn[i]] == 'dead':
                    print('\n' + str(monster[playturn[i]]) + ' still dead')
                    break
                elif monster[playturn[i]] == 'flew':
                    print('\n' + str(monster[playturn[i]]) + ' still running')
                    break
                print('\n' + str(monster[playturn[i]]) + ' turn')
                # print(playturn[i], ActionTimer, playturn, monsterdata)
                # print(monsterdata[playturn[i]])
                varhp = monsterdata[playturn[i]][0][0]
                varatk = monsterdata[playturn[i]][7][0]
                varmagic = monsterdata[playturn[i]][9][0]
                varspd = monsterdata[playturn[i]][11][0]
                varhit = monsterdata[playturn[i]][13][0]
                vartech = monsterdata[playturn[i]][17]
                varcounter = monsterdata[playturn[i]][18]

                # print(varhp, varatk, varmagic, varspd, varhit, vartech, varcounter)
                deathtest = False
                for deathcounter in range(128):
                    choosenumber = randint(0, 2)
                    choosenone = party[choosenumber]
                    templayer = ReadPlayerData(choosenone)
                    # print(templayer)
                    # print(templayer[1])
                    if int(templayer[1]) > 0:
                        # print(choosenone, 'is alive')
                        if choosenumber == 0:
                            columletter = temprpp01
                        elif choosenumber == 1:
                            columletter = temprpp11
                        elif choosenumber == 2:
                            columletter = temprpp21

                        decideaction = randint(0, 100)
                        # print(monsterdata[playturn[i]])
                        # print(monsterdata[playturn[i]][17])

                        if decideaction == 100:
                            print(str(monster[playturn[i]]), 'ran away')
                            monsterdata[playturn[i]][0][0] = 0
                            monster[playturn[i]] = 'flew'
                            deathtest = True
                            break
                        elif 0 <= decideaction < 66:
                            # Basic Attack Below
                            print(str(monster[playturn[i]]), 'attacked', choosenone, 'dealing: ',
                                  math.floor((int(varatk) * (256 - int(templayer[4])) / 256)))
                            # print(type(Characters[get_column_letter(columletter + 2) + '4'].value), type(get_column_letter(columletter+2)), type(varatk), type(templayer[4]))
                            Characters[get_column_letter(columletter + 2) + '4'] = int(
                                Characters[get_column_letter(columletter + 2) + '4'].value) - math.floor(
                                int(varatk) * (256 - int(templayer[4])) / 256)
                            if Characters[get_column_letter(columletter + 2) + '4'].value <= 0:
                                Characters[get_column_letter(columletter + 2) + '4'] = 0
                            playerdatavolatileobj.save(playerdatavolatile)
                            deathtest = True
                            break
                        else:
                            # Tech Attack Below
                            # print('do something')
                            techchoice = randint(0, len(monsterdata[playturn[i]][17]) - 1)
                            thistech = monsterdata[playturn[i]][17][techchoice]
                            # print(thistech)
                            # print(monsterdata[playturn[i]][0])
                            # print(monsterdata[playturn[i]][7])
                            # print(monsterdata[playturn[i]][9])
                            # print(monsterdata[playturn[i]][11])
                            # print(monsterdata[playturn[i]][13])
                            techdamage = tech(thistech, HP=int(monsterdata[playturn[i]][0][0]),
                                              ATK=int(monsterdata[playturn[i]][7][0]),
                                              MAG=int(monsterdata[playturn[i]][9][0]),
                                              SPD=int(monsterdata[playturn[i]][11][0]),
                                              HIT=int(monsterdata[playturn[i]][13][0]), RNG=1)
                            # print(techdamage)

                            if grouptechs(thistech):
                                for j in range(3):
                                    if j == 0:
                                        columletter = temprpp01
                                    elif j == 1:
                                        columletter = temprpp11
                                    elif j == 2:
                                        columletter = temprpp21
                                    choosenone = party[j]
                                    if techdamage[1] == 'PHY':
                                        reducedamage = math.floor(
                                            techdamage[0] * (256 - int(
                                                Characters[get_column_letter(columletter + 1) + '7'].value)) / 256)
                                        print(str(monster[playturn[i]]), 'used', thistech, 'in', choosenone,
                                              'dealing: ', reducedamage)
                                        Characters[get_column_letter(columletter + 2) + '4'] = int(
                                            Characters[get_column_letter(columletter + 2) + '4'].value) - reducedamage
                                    elif techdamage[1] == 'LIG' or techdamage[1] == 'SHA' or techdamage[1] == 'FIR' or \
                                            techdamage[1] == 'WAT':
                                        reducedamage = math.floor(
                                            techdamage[0] * (101 - int(
                                                Characters[get_column_letter(columletter + 1) + '15'].value)) / 101)
                                        print(str(monster[playturn[i]]), 'used', thistech, 'in', choosenone,
                                              'dealing: ', reducedamage)
                                        Characters[get_column_letter(columletter + 2) + '4'] = int(
                                            Characters[get_column_letter(columletter + 2) + '4'].value) - reducedamage

                                    if Characters[get_column_letter(columletter + 2) + '4'].value <= 0:
                                        Characters[get_column_letter(columletter + 2) + '4'] = 0
                            else:
                                reducedamage = 0
                                if techdamage[1] == 'PHY':
                                    reducedamage = math.floor(
                                        techdamage[0] * (256 - int(
                                            Characters[get_column_letter(columletter + 1) + '7'].value)) / 256)
                                    Characters[get_column_letter(columletter + 2) + '4'] = int(
                                        Characters[get_column_letter(columletter + 2) + '4'].value) - reducedamage
                                elif techdamage[1] == 'LIG' or techdamage[1] == 'SHA' or techdamage[1] == 'FIR' or \
                                        techdamage[1] == 'WAT':
                                    reducedamage = math.floor(
                                        techdamage[0] * (101 - int(
                                            Characters[get_column_letter(columletter + 1) + '15'].value)) / 101)
                                    Characters[get_column_letter(columletter + 2) + '4'] = int(
                                        Characters[get_column_letter(columletter + 2) + '4'].value) - reducedamage
                                print(str(monster[playturn[i]]), 'used', thistech, 'in', choosenone, 'dealing: ',
                                      reducedamage)
                            if Characters[get_column_letter(columletter + 2) + '4'].value <= 0:
                                Characters[get_column_letter(columletter + 2) + '4'] = 0

                            playerdatavolatileobj.save(playerdatavolatile)
                            deathtest = True
                            break

                if deathtest is False:
                    print('You lost')
                    lose = True
                break
    print('[Process:', localvar, ']', 'Success')


globalvar = 0
maxspeed = 20
subglobalval = globalvar
print('[Process:', subglobalval, ']', 'Starting')
globalvar = globalvar + 1

party = ['crono', 'marle', 'lucca']
# ResetGame()
# NewGamePlayer('crono')
# NewGamePlayer('marle')
# NewGamePlayer('lucca')
# EquipItem('Armor', "Moonbeam Armor", 'crono')
LevelUpPlayer('crono', 10)
LevelUpPlayer('marle', 10)
LevelUpPlayer('lucca', 10)
# print(ReadMonsterData('cyrus'))
# UseItem('Speed Tab', 'crono', 'crono')
# UseItem('Power Tab', 'crono', 'crono')
# UseItem('Magic Tab', 'crono', 'crono')

# ReadPlayerData('crono')

# Inventory('add', 'tonic', 'item')
# Inventory('add', 'tonic', 'item')
# Inventory('add', 'tonic', 'item')
# Inventory('add', 'tonic', 'item')
# Inventory('add', 'tonic', 'item')
# Inventory('add', 'revive', 'item')
# Inventory('add', 'revive', 'item')
# Inventory('add', '10', 'gold')
# learnTech('cyclone', 'crono')
battle('prison towers - present', mod='vision')
# battle('yakra - middle age', mod='vision')

print('[Process:', subglobalval, ']', 'End')
